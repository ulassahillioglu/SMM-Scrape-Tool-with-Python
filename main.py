import random
from selenium.webdriver import Chrome
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver import Firefox
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
import pandas as pd
import html.parser
import time
from openpyxl import Workbook
from itertools import chain

followers_list = [] 
global b, c, d, e, f, m, n, j, website
b = input("Word to Search 1: ")
c = input("Word to Search 2 (switches with alternative word if desired): ")  ##TR tedarikçilerde aynı servisler farklı isimlere sahip olabliyor
                                                                              ## Kadın > Bayan, Özel > Belirlenebilir              
f = input("word to Search 3: ")
# k = input("Alternative Word: ") uncomment if necessary and add to functions with or operator
ml = float(input("Max Price: "))
d = "detay"
e = input("Enter file name : ")
e = str(e) + ".xlsx"
m = "açıklama"
n = "mevcuttur"
j = "notlar"

def ScrapeTool(a): #Left to right 'provider - desired keyword1 - desired keyword1 - omitted word - file name'
    
    template = []
    suppId =[]
    products = []
    prices = [] 
    minAndmax = []

    profile_path = "C:/Users/"   ##Enter your Firefox profile if you wish to use geckodriver
    options=Options()
    options.set_preference('profile', profile_path)
    options.accept_insecure_certs  #SSL Cert bypass
    service = Service('chromedriver.exe')

    driver = webdriver.Chrome()
    driver.implicitly_wait(10)
    driver.get(a)
    WebDriverWait(driver,random.randint(5,15)) 
    try:
        info = driver.page_source
        soup = BeautifulSoup(info,features="html.parser")
        job_elements = soup.find_all("td")
        for job_element in job_elements:
                
                new_job_element = job_element.text.strip().replace("\t","").replace("\n","").replace("\r","") #finds table elements
                template.append(new_job_element)
    except Exception as excpt:
        print(excpt)
    # print(template)

    for element in template:
        x = template.index(element) 
        #filters table elements as requested, uncomment first if block and comment second if block to use with alternative word
        # if b in element.lower() and c in element.lower() and f in element.lower() and d not in element.lower() and m not in element.lower() and n not in element.lower() and j not in element.lower() or b in element.lower() and k in element.lower() and f in element.lower() and d not in element.lower() and m not in element.lower() and n not in element.lower() and j not in element.lower(): 
        if b in element.lower() and c in element.lower() and f in element.lower() and d not in element.lower() and m not in element.lower() and n not in element.lower() and j not in element.lower(): 
                try:
                    if float(template[x+1].replace(",",".").replace("TL","").replace("₺","").strip()) <=ml:
                        template[x-1] = str(template[x-1]).strip()
                        suppId.append(int(template[x-1]))
                        
                        template[x] = str(template[x]).strip()
                        products.append(template[x])
                        # print(template[ele])
                    
                        template[x+1] = template[x+1].replace("$","").replace("₺","").replace(" ","").replace("TL","")
                        
                        prices.append(float(template[x+1]))
                        # print(template[ele+1])

                        
                        minAndmax.append([int(template[x+2].strip()),int(template[x+3].strip())])
                        
                except TypeError:
                    continue
                except ValueError:
                    pass
        
    driver.close()

    
    # print(suppId)
    # print(products)
    # print(prices)
    # print(minAndmax)
    # print(minAndmax)

    print(len(suppId),len(products),len(prices),len(minAndmax))
    
    scrape_dict = { ## first we create a dictionary, later we will convert it to Pandas dataframe
        "ID" : suppId,
        "Product" : products,
        "Price" : prices,
        "Min and Max Amount" : minAndmax,
    }
   
    df = pd.DataFrame(scrape_dict)
    global book
    try:
        book = load_workbook(e) ##create a new file for new tests if the file with given name does not exist
    except:
        wb = Workbook()
        wb.save(e)
        book = load_workbook(e)
    writer = pd.ExcelWriter(e, engine='openpyxl',mode="a",if_sheet_exists="replace") 
    writer.book = book
    nameToAppend = a.split("/")
    pageName = nameToAppend[2]
    df.to_excel(writer, sheet_name= pageName,index=False)
    writer.save()
    writer.close()
    print("Moving to next one")

      
trProvider = ["https://sosyalbayiniz.net/services","https://paneliniz.com/services","https://medyabayim.com/services","https://smmturk.net/services","https://medyapanelim.com/services","https://igresellers.com/services",
"https://smmfull.com/services","https://smpanel.net/services","https://sosyalatom.com/services","https://panelhizmetleri.net/services"]


for provider in (trProvider):
    if provider == "https://sosyalatom.com/services":
        template = []
        suppId =[]
        products = []
        prices = [] 
        minnb = []
        maxnb = []
        url = "https://sosyalatom.com/services"



        profile_path = "C:/Users/"
        options=Options()
        options.set_preference('profile', profile_path)
        options.accept_insecure_certs  #SSL Cert bypass
        service = Service('chromedriver.exe')

        driver = webdriver.Chrome()
        driver.implicitly_wait(10)
        driver.get(url)
        website = driver.page_source.encode("utf-8")
        WebDriverWait(driver,random.randint(5,15))

        soup = bs(website,"html.parser")
        idlist = soup.select("div.font-medium > span",{"title":'data-filter-table-service-id'})
        servicelist = soup.select("div>div>div > span",{'title':"sa_muted"})
        # pricelist = soup.select("td[data-title = '1K Ücreti']")
        # minamount = soup.select("td[data-title = 'Min Sipariş']")
        # maxamount = soup.select("td[data-title = 'Max Sipariş']")



        driver.close()

        for data in idlist:
            data = data.text.replace("\t","").replace("\n","").strip()
            suppId.append(data)
        for service in servicelist:
            products.append(service.text.replace("\t","").strip())

        # print(suppId)

        pricelist = []
        servicenames = []
        serviceId= []
        for ele in suppId:
            try:
                indnumber = int(suppId.index(ele))
                s = suppId[indnumber].split("-")
                pricelist.append(s[1].replace("TL","").strip())
                servicenames.append(s[0])
                serviceId.append(suppId[indnumber-1])
            except:
                pass
        i = 0
        supplier = []
        while i<=len(suppId)-2:
            supplier.append([suppId[i],suppId[i+1]])
            i = i+2
        print(len(products))
        i = 0
        newlist= []
        while i<=len(products)-2:
            newlist.append(str(products[i]) +str(products[i+1]))
            i = i+2
        print(len(newlist))

        minlist = []
        maxlist= []
        i=0
        while i<=len(newlist)-2:
            minlist.append(newlist[i])
            maxlist.append(newlist[i+1])
            i = i+2
        print(len(minlist),len(maxlist))

        services = list(zip(serviceId,servicenames,pricelist,minlist,maxlist))

        # print(services)


        template = list(chain.from_iterable(services))
        # print(template)
        suppId =[]
        products = []
        prices = [] 
        minAndmax = []
        for service in template:
            y = template.index(service)
            # if b in service.replace("İ","I").lower() and c in service.replace("İ","I").lower() and f in service.replace("İ","I").lower() or b in service.replace("İ","I").lower() and k in service.replace("İ","I").lower() and f in service.replace("İ","I").lower():
            if b in service.replace("İ","I").lower() and c in service.replace("İ","I").lower() and f in service.replace("İ","I").lower():   

                    suppId.append(int(template[y-1].strip()))
                    products.append(template[y].strip())
                    prices.append(float(template[y+1].strip().replace("TL","").replace(",",".")))
                    minAndmax.append([template[y+2].replace("\t",""),template[y+3]])



        scrape_dict = {
            "ID" : suppId,
            "Product" : products,
            "Price" : prices,
            "Min and Max Amount" : minAndmax,
        }

        # print(minAndmax)
        df = pd.DataFrame(scrape_dict)
        
        try:
            book = load_workbook(e) ##create a new file for new tests
        except:
            wb = Workbook()
            wb.save(e)
            book = load_workbook(e)
        writer = pd.ExcelWriter(e, engine='openpyxl',mode="a",if_sheet_exists="new") 
        writer.book = book
        pageName = "sosyalatom " + b[0:3] + " " + c 
        df.to_excel(writer, sheet_name= pageName,index=False, startrow=0)
        writer.save()
        writer.close()
        print(len(suppId),len(products),len(prices),len(minAndmax))
        print("Moving to next one")
    elif provider == "https://panelhizmetleri.net/services":
        
        template = []
        suppId =[]
        products = []
        prices = [] 
        minAndmax = []
        url = "https://panelhizmetleri.net/services"

        profile_path = "C:/Users/"
        options=Options()
        options.set_preference('profile', profile_path)
        options.accept_insecure_certs  #SSL Cert bypass
        service = Service('chromedriver.exe')

        driver = webdriver.Chrome()
        driver.implicitly_wait(10)
        driver.get(url)
        website = driver.page_source.encode("utf-8")
        WebDriverWait(driver,random.randint(5,15))
        soup = BeautifulSoup(website,"html.parser")
        
        data2 = soup.find_all("td")
        driver.close()
        for element in data2:
            new_job_element = element.text.strip().replace("\t","").replace("\n","").replace("\r","") #finds table elements
            template.append(new_job_element)
        for ele in template:
            z = template.index(ele)
            #uncomment to use with alternative word
            # if b in ele.lower() and c in ele.lower() and f in ele.lower() and d not in ele.lower() and m not in ele.lower() and n not in ele.lower() or b in ele.lower() and k  in ele.lower() and f in ele.lower() and d not in ele.lower() and m not in ele.lower() and n not in ele.lower(): 
            if b in ele.lower() and c in ele.lower() and f in ele.lower() and d not in ele.lower() and m not in ele.lower() and n not in ele.lower():
                template[z+1] = template[z+1].replace("$","").replace("₺","").replace(" ","").replace("TL","")
                if float(template[z+1].replace(",","."))<=ml:
                    try:
                        template[z-1] = str(template[z-1]).strip()
                        suppId.append(int(template[z-1]))
                        
                        template[z] = str(template[z]).strip()
                        products.append(template[z])
                        # print(template[ele])
                    
                        template[z+1] = template[z+1].replace("$","").replace("₺","").replace(" ","").replace("TL","")
                        
                        prices.append(float(template[z+1]))
                        # print(template[ele+1])

                        template[z+3] = template[z+3].strip()
                        template[z+4] = template[z+4].strip()
                        minAndmax.append([int(template[z+3]),int(template[z+4])])
                    except Exception as excpt:
                        print(excpt)
        scrape_dict = {
            "ID" : suppId,
            "Product" : products,
            "Price" : prices,
            "Min and Max Amount" : minAndmax,
        }
    
        df = pd.DataFrame(scrape_dict)
        
        try:
            book = load_workbook(e) ##create a new file for new tests
        except:
            wb = Workbook()
            wb.save(e)
            book = load_workbook(e)
        writer = pd.ExcelWriter(e, engine='openpyxl',mode="a",if_sheet_exists="new") 
        writer.book = book
        pageName = "panelhizmetleri"
        df.to_excel(writer, sheet_name= pageName,index=False)
        writer.save()
        writer.close()
        print(len(suppId),len(products),len(prices),len(minAndmax))
        print("Moving to next one")
   
   

    elif provider == "https://smpanel.net/services":
        
        from bs4 import BeautifulSoup as bs
        template = []
        suppId =[]
        products = []
        prices = [] 
        minAndmax = []
        url = "https://smpanel.net/services"


        profile_path = "C:/Users/"
        options=Options()
        options.set_preference('profile', profile_path)
        options.accept_insecure_certs  #SSL Cert bypass
        service = Service('chromedriver.exe')

        driver = webdriver.Chrome()
        driver.implicitly_wait(10)
        driver.get(url)
        website = driver.page_source.encode("utf-8")
        
        soup = bs(website,"html.parser")
        idlist = soup.select("td[data-title = 'ID']")
        servicelist = soup.select("td[data-title = 'Servis']")
        pricelist = soup.select("td[data-title = '1000 Adet']")
        minmaxamount = soup.select("td[data-title = 'Min-Maks']")

        for data in idlist:
            data = data.text
            suppId.append(data)
        for service in servicelist:
            products.append(service.text)
        for price in pricelist:
            prices.append(price.text)
        for amo in minmaxamount:
            minAndmax.append(amo.text)

        zipped = list(zip(suppId,products,prices,minAndmax))



        template = list(chain.from_iterable(zipped))
        # print(template)
        suppId =[]
        products = []
        prices = [] 
        minAndmax = []
        for product in template:
                v = template.index(product)
                # if b in product.lower() and c in product.lower() and f in product.lower() or b in product.lower() and k in product.lower() and f in product.lower() :
                if b in product.lower() and c in product.lower() and f in product.lower():
                    try:
                        if float(template[v+1].replace(",","."))<=ml:
                            suppId.append(int(template[v-1]))
                            products.append(template[v])
                            prices.append(float(template[v+1].replace(",",".")))
                            minAndmax.append(template[v+2])
                    except:
                        pass



        scrape_dict = {
                "ID" : suppId,
                "Product" : products,
                "Price" : prices,
                "Min and Max Amount" : minAndmax,
            }

        # print(minAndmax)
        df = pd.DataFrame(scrape_dict)

        try:
            book = load_workbook(e) ##create a new file for new tests
        except:
            wb = Workbook()
            wb.save(e)
            book = load_workbook(e)
        writer = pd.ExcelWriter(e, engine='openpyxl',mode="a",if_sheet_exists="new") 
        writer.book = book
        pageName = "smpanel"
        df.to_excel(writer, sheet_name= pageName,index=False)
        writer.save()
        writer.close()
        print(len(suppId),len(products),len(prices),len(minAndmax))
        print("Moving to next one")

    elif provider == "https://igresellers.com/services":
        from bs4 import BeautifulSoup as bs
        template = []
        suppId =[]
        products = []
        prices = [] 
        minnb = []
        maxnb = []
        url = "https://igresellers.com/services"



        profile_path = "C:/Users/"
        options=Options()
        options.set_preference('profile', profile_path)
        options.accept_insecure_certs  #SSL Cert bypass
        service = Service('chromedriver.exe')

        driver = webdriver.Chrome()
        driver.implicitly_wait(10)
        driver.get(url)
        website = driver.page_source.encode("utf-8")
        WebDriverWait(driver,random.randint(5,15))

        soup = bs(website,"html.parser")
        idlist = soup.select("td[data-title = 'ID']")
        servicelist = soup.select("td[data-title = 'Servis']")
        pricelist = soup.select("td[data-title = '1K Ücreti']")
        minamount = soup.select("td[data-title = 'Min Sipariş']")
        maxamount = soup.select("td[data-title = 'Max Sipariş']")



        driver.close()

        for data in idlist:
            data = data.text.replace("\t","").strip()
            suppId.append(data)
        for service in servicelist:
            products.append(service.text.replace("\t","").strip())
        for price in pricelist:
            prices.append(price.text.replace("\t","").strip())
        for amo in minamount:
            minnb.append(amo.text.replace("\t","").strip())
        for amou in maxamount:
            maxnb.append(amou.text.replace("\t","").strip())


        zipped = list(zip(suppId,products,prices,minnb,maxnb))



        template = list(chain.from_iterable(zipped))
        # print(template)
        suppId =[]
        products = []
        prices = [] 
        minAndmax = []
        for ele in template:
            i = template.index(ele)
            
            # if b in ele.replace("İ","I").lower() and c in ele.replace("İ","I").lower() and f in ele.replace("İ","I").lower() or b in ele.replace("İ","I").lower() and k in ele.replace("İ","I").lower() and f in ele.replace("İ","I").lower() :
            if b in ele.replace("İ","I").lower() and c in ele.replace("İ","I").lower() and f in ele.replace("İ","I").lower():
                try:
                    if float(template[i+1].replace("TL","").strip())<= ml:
                        suppId.append(int(template[i-1].strip()))
                        products.append(template[i].strip())
                        prices.append(float(template[i+1].strip().replace("TL","")))
                        minAndmax.append([template[i+2].replace("\t",""),template[i+3]])
                except:
                    pass



        scrape_dict = {
            "ID" : suppId,
            "Product" : products,
            "Price" : prices,
            "Min and Max Amount" : minAndmax,
        }

        # print(minAndmax)
        df = pd.DataFrame(scrape_dict)

        try:
            book = load_workbook(e) ##create a new file for new tests
        except:
            wb = Workbook()
            wb.save(e)
            book = load_workbook(e)
        writer = pd.ExcelWriter(e, engine='openpyxl',mode="a",if_sheet_exists="new") 
        writer.book = book
        pageName = "igresellers " + b[0:3] + " " + c 
        df.to_excel(writer, sheet_name= pageName,index=False, startrow=0)
        writer.save()
        writer.close()
        print(len(suppId),len(products),len(prices),len(minAndmax))
        print("Moving to next one")
    elif provider == "https://smmturk.net/services":
        url = "https://smmturk.net/services"
        template = list()
        profile_path = "C:/"
        options=Options()
        options.set_preference('profile', profile_path)
        options.accept_insecure_certs  #SSL Cert bypass

        driver = webdriver.Chrome()
        driver.get(url)

        info = driver.page_source
        soup = BeautifulSoup(info,'html.parser')
        elements = soup.select("table p")
        driver.close()

        data = list()

        for element in elements:
            data.append(element.text)
        servicenames = list()
        serviceID = list()
        for ele in data:
            try:
                indnumber = int(data.index(ele))
                s = data[indnumber].split("-")
                serviceID.append(s[0].strip())
                servicenames.append(s[1])
                
            except:
                pass


        pricestemplate = soup.select("td.mobile-tr")
        prices = [price.text for price in pricestemplate]
        pricelist = list()
        minmax = list()
        print(len(prices))
        i = 0
        while i<= len(prices)-2:
            pricelist.append(prices[i])
            minmax.append(prices[i+1])
            i = i + 2

        zipped = list(zip(serviceID,servicenames,pricelist,minmax))

        # print(zipped)


        template = list(chain.from_iterable(zipped))

        suppId = list()
        products = list()
        prices = list()
        minAndmax = list()
        for element in template:
            x = int(template.index(element))
            # if b in element.lower() and  c in element.lower() and f in element.lower() or b  in element.lower() and k in element.lower() and f in element.lower():
            if b in element.lower() and  c in element.lower() and f in element.lower():
                
                try:
                    if float(template[x+1].replace("TL","").strip()) <= ml:
                        
                        suppId.append(template[x-1])
                        products.append(element)
                        prices.append(float(template[x+1].replace("TL","")))
                        minAndmax.append(template[x+2])
                except:
                    pass

        scrape_dict = {
                    "ID" : suppId,
                    "Product" : products,
                    "Price" : prices,
                    "Min and Max Amount" : minAndmax,
                }

        # print(minAndmax)
        df = pd.DataFrame(scrape_dict)

        try:
            book = load_workbook(e) ##create a new file for new tests
        except:
            wb = Workbook()
            wb.save(e)
            book = load_workbook(e)
        writer = pd.ExcelWriter(e, engine='openpyxl',mode="a",if_sheet_exists="new") 
        writer.book = book
        pageName = "smmturk " + b[0:3] + " " + c 
        df.to_excel(writer, sheet_name= pageName,index=False, startrow=0)
        writer.save()
        writer.close()
        print(len(suppId),len(products),len(prices),len(minAndmax))
        print("Moving to next one")
    else:
        ScrapeTool(provider)
        time.sleep(15)
        print("Taking a break")

print("Completed")   
